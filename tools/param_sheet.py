"""参数单导出：工艺参数 → Excel（人读输出拍板 2026-07-02：参数用 Excel，建议用中文文本）。

定位：给老师傅复核/操作的**人读参数单**（与 D2 复核 Excel 表流同介质），不是 PLC 下发报文
（下发通道待清单 B4/B9）。安全状态如实呈现：未过闸门的单子醒目标注"禁止照此操作"并列 violations。
openpyxl 不可用时降级写 CSV（utf-8-sig，Excel 打开中文不乱码），返回实际落盘路径。

用法：& "D:\\Glass Agents\\.venv\\Scripts\\python.exe" tools\\param_sheet.py 参数.json ^
        [--baseline 基准.json] [--furnace-id F1] [-o 输出.xlsx]
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # 支持直接 `python tools/param_sheet.py`

from schemas.process_params import ProcessParams  # noqa: E402

if TYPE_CHECKING:
    from tools.constraints import CheckResult

_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_OUT_DIR = _ROOT / "data" / "outbox"

# 标量参数：字段名 → (中文名, 单位)；与 param_translator 骨架用词一致，不造第二套叫法
_SCALAR_FIELDS: tuple[tuple[str, str, str], ...] = (
    ("temp_upper", "上炉温", "℃"),
    ("temp_lower", "下炉温", "℃"),
    ("convection_speed", "对流风速", "-"),
    ("convection_ratio_upper_lower", "上下对流配比", "-"),
    ("oscillation_speed", "摆动速度", "-"),
    ("oscillation_amplitude", "摆动幅度", "-"),
    ("heating_duration_s", "加热时长", "s"),
)
_GLASS_ZH = {"ultra_clear": "超白", "clear": "普白"}
_MODE_ZH = {"high_quality": "高质量", "high_efficiency": "高效率"}


def _fmt(v: float | None) -> str:
    """数值格式化：缺 → 空串（缺值不猜）。"""
    return "" if v is None else f"{v:g}"


def build_rows(
    params: ProcessParams,
    check: CheckResult,
    *,
    baseline: ProcessParams | None = None,
    meta: dict | None = None,
) -> list[list[str]]:
    """组参数单行（xlsx 与 CSV 共用同一份行数据，格式只做一次）。"""
    m = meta or {}
    gate = "✅ 通过" if check.within_limits else "⛔ 未通过——禁止照此操作"
    rows: list[list[str]] = [
        ["钢化炉工艺参数单", "", "", "", "", ""],
        ["炉号", str(m.get("furnace_id", "")), "样本号", str(m.get("sample_id", "")), "", ""],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M"), "安全闸门", gate, "", ""],
        ["", "", "", "", "", ""],
        ["参数", "字段名", "当前值", "单位", "基准值", "Δ"],
        [
            "规格",
            "spec",
            f"{_GLASS_ZH.get(params.glass_type, params.glass_type)} {params.thickness_mm:g}mm "
            f"{_MODE_ZH.get(params.quality_mode, params.quality_mode)}",
            "",
            "",
            "",
        ],
    ]
    for name, zh, unit in _SCALAR_FIELDS:
        cur = float(getattr(params, name))
        base = float(getattr(baseline, name)) if baseline is not None else None
        delta = None if base is None else cur - base
        rows.append([zh, name, _fmt(cur), unit, _fmt(base), "" if delta is None else f"{delta:+g}"])
    for i, (t, role) in enumerate(zip(params.zone_temps, params.zone_roles), start=1):
        base_t = None
        if baseline is not None and i - 1 < len(baseline.zone_temps):
            base_t = float(baseline.zone_temps[i - 1])
        delta = None if base_t is None else float(t) - base_t
        rows.append(
            [f"分区{i}温度({role})", f"zone_temps[{i - 1}]", _fmt(float(t)), "℃", _fmt(base_t),
             "" if delta is None else f"{delta:+g}"]
        )
    if not check.within_limits:
        rows.append(["", "", "", "", "", ""])
        rows.append(["⛔ 未通过安全闸门，禁止照此操作。问题清单：", "", "", "", "", ""])
        for v in check.violations:
            rows.append([v, "", "", "", "", ""])
    return rows


def write_param_sheet(
    params: ProcessParams,
    check: CheckResult,
    *,
    baseline: ProcessParams | None = None,
    meta: dict | None = None,
    out_path: Path,
) -> Path:
    """写参数单：优先 xlsx（冻结表头/闸门行醒目），openpyxl 不可用降级 CSV。返回实际路径。"""
    rows = build_rows(params, check, baseline=baseline, meta=meta)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        csv_path = out_path.with_suffix(".csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)
        return csv_path

    wb = Workbook()
    ws = wb.active
    ws.title = "参数单"
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    gate_cell = ws["D3"]
    gate_cell.font = Font(bold=True, color="006100" if check.within_limits else "9C0006")
    ws.freeze_panes = "A6"                       # 冻结头部区块+表头
    for col, width in zip("ABCDEF", (22, 30, 14, 8, 12, 10)):
        ws.column_dimensions[col].width = width
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")
    wb.save(out_path)
    return out_path


def default_out_path(out_dir: Path | None = None) -> Path:
    """默认落盘路径：data/outbox/参数单_<时间戳>.xlsx（outbox 已 gitignore）。"""
    d = Path(out_dir) if out_dir is not None else _DEFAULT_OUT_DIR
    return d / f"参数单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def main() -> int:
    """CLI：读参数 JSON →（现算）闸门 → 出参数单。"""
    from tools.constraints import validate  # 延迟导入：读 config 的闸门只在真出单时用

    ap = argparse.ArgumentParser(description="工艺参数 → Excel 参数单（未过闸门如实标注禁止照做）")
    ap.add_argument("params_json", type=Path, help="ProcessParams JSON")
    ap.add_argument("--baseline", type=Path, default=None, help="基准配方 JSON（出 Δ 列）")
    ap.add_argument("--furnace-id", default="", help="炉号（表头展示）")
    ap.add_argument("-o", "--out", type=Path, default=None, help="输出路径（默认 data/outbox/参数单_时间戳.xlsx）")
    args = ap.parse_args()

    params = ProcessParams.model_validate_json(args.params_json.read_text(encoding="utf-8"))
    baseline = None
    if args.baseline is not None:
        baseline = ProcessParams.model_validate_json(args.baseline.read_text(encoding="utf-8"))
    prev = baseline.to_param_set() if baseline is not None else None
    check = validate(params.to_param_set(), prev=prev)

    out = write_param_sheet(
        params, check, baseline=baseline,
        meta={"furnace_id": args.furnace_id},
        out_path=args.out if args.out is not None else default_out_path(),
    )
    print(f"参数单已写 {out}（安全闸门: {'通过' if check.within_limits else '未通过，单内已标注禁止照做'}）")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except Exception:
        pass
    sys.exit(main())
