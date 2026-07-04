"""生成老师傅标注表模板：CSV(utf-8-sig，Excel 中文不乱码) + xlsx(枚举下拉、冻结表头、字段字典页)。

列定义在本文件 COLUMNS 单一来源，与 schemas.ArchiveSample / ProcessParams / CONVENTIONS 对齐。
示例行仅为**格式示意**（sample_id 前缀“示例-”），非工厂真值；现场用真实基准配方替换 baseline_*。

呈现减负（2026-07-03，列契约 39 列不动）：第 2 行为中文说明行（sample_id 带“示例-”前缀，
导入器按既有机制跳过，勿删）；xlsx 里老师傅要填的列黄底标 ★，其余灰底且默认折叠
（记录员点列上方“+”展开），出炉后补列浅蓝——让师傅打开只看见要填的。

用法：& "D:\\Glass Agents\\.venv\\Scripts\\python.exe" tools\\make_annotation_template.py
产物：data/annotation/标注表模板.csv 与（openpyxl 可用时）data/annotation/标注表模板.xlsx
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "annotation"

# 枚举列的允许值（xlsx 下拉 + 字典）
ENUMS: dict[str, list[str]] = {
    "glass_type": ["ultra_clear", "clear"],
    "quality_mode": ["high_quality", "high_efficiency"],
    "expert_quality_grade": ["A", "B", "C"],
    "measured_quality_grade": ["A", "B", "C"],
    "ccp_is_calibrated": ["TRUE", "FALSE"],
    "is_ground_truth": ["TRUE", "FALSE"],
}

# 每列：(字段名, 组, 含义/单位/允许值, 谁填, [示例1, 示例2])
COLUMNS: list[tuple[str, str, str, str, list[str]]] = [
    ("sample_id", "A 标识", "唯一样本号", "记录员", ["示例-001", "示例-002"]),
    ("created_at", "A 标识", "ISO 时间 2026-07-02T09:30", "系统", ["2026-07-02T09:30", "2026-07-02T10:05"]),
    ("operator_id", "A 标识", "老师傅工号（测标签一致性）", "记录员", ["S01", "S01"]),
    ("source", "A 标识", "产线/班次", "记录员", ["line1/早班", "line1/早班"]),
    ("thickness_mm", "B 规格·分桶键", "厚度 mm", "记录员", ["8", "8"]),
    ("glass_type", "B 规格·分桶键", "品类 {ultra_clear,clear}", "记录员", ["clear", "ultra_clear"]),
    ("quality_mode", "B 规格·分桶键", "模式 {high_quality,high_efficiency}", "记录员", ["high_quality", "high_efficiency"]),  # noqa: E501
    ("x0_95_nm", "C 输入·状况X", "应力斑 X0.95，nm（可得则填）", "系统", ["82", "70"]),
    ("iso_t_pct", "C 输入·状况X", "IsoT %（暂无→留空/TODO(plant)）", "系统", ["", ""]),
    ("ccp_value", "C 输入·状况X", "CCP（暂无→留空/TODO(plant)）", "系统", ["", ""]),
    ("ccp_is_calibrated", "C 输入·状况X", "CCP 是否标定 TRUE/FALSE", "系统", ["FALSE", "FALSE"]),
    ("stress_image_path", "C 输入·状况X", "应力斑图路径（有则填）", "系统", ["", ""]),
    ("stress_image_sha256", "C 输入·状况X", "图像 sha256（有则填）", "系统", ["", ""]),
    ("condition_note", "C 输入·状况X", "工况备注（结构化字段待定→先文本）", "记录员", ["炉温稳定", "换批次首片"]),
    ("baseline_zone_temps_c", "D 基准(只读参考)", "基准分区温度 ℃，分号隔", "系统预填", ["700;705;700", "700;705;700"]),
    ("baseline_zone_roles", "D 基准(只读参考)", "基准分区角色，分号隔 {center,edge,...}", "系统预填", ["center;center;edge", "center;center;edge"]),  # noqa: E501
    ("baseline_temp_upper_c", "D 基准(只读参考)", "基准上炉温 ℃", "系统预填", ["700", "700"]),
    ("baseline_temp_lower_c", "D 基准(只读参考)", "基准下炉温 ℃", "系统预填", ["690", "690"]),
    ("baseline_convection_speed", "D 基准(只读参考)", "基准对流风速", "系统预填", ["1.0", "1.0"]),
    ("baseline_convection_ratio_upper_lower", "D 基准(只读参考)", "基准上下对流配比", "系统预填", ["1.0", "1.0"]),
    ("baseline_oscillation_speed", "D 基准(只读参考)", "基准摆动速度", "系统预填", ["1.0", "1.0"]),
    ("baseline_oscillation_amplitude", "D 基准(只读参考)", "基准摆动幅度", "系统预填", ["1.0", "1.0"]),
    ("baseline_heating_duration_s", "D 基准(只读参考)", "基准加热时长 s", "系统预填", ["220", "220"]),
    ("final_zone_temps_c", "E 老师傅最终参数", "最终分区温度 ℃，分号隔（预填=基准，改需动的）", "老师傅", ["705;708;700", "700;705;700"]),  # noqa: E501
    ("final_zone_roles", "E 老师傅最终参数", "最终分区角色，分号隔（与温度等长）", "老师傅", ["center;center;edge", "center;center;edge"]),  # noqa: E501
    ("final_temp_upper_c", "E 老师傅最终参数", "最终上炉温 ℃", "老师傅", ["705", "700"]),
    ("final_temp_lower_c", "E 老师傅最终参数", "最终下炉温 ℃", "老师傅", ["690", "690"]),
    ("final_convection_speed", "E 老师傅最终参数", "最终对流风速", "老师傅", ["1.2", "1.0"]),
    ("final_convection_ratio_upper_lower", "E 老师傅最终参数", "最终上下对流配比", "老师傅", ["1.0", "1.0"]),
    ("final_oscillation_speed", "E 老师傅最终参数", "最终摆动速度", "老师傅", ["1.0", "1.0"]),
    ("final_oscillation_amplitude", "E 老师傅最终参数", "最终摆动幅度", "老师傅", ["1.0", "1.0"]),
    ("final_heating_duration_s", "E 老师傅最终参数", "最终加热时长 s", "老师傅", ["235", "220"]),
    ("expert_quality_grade", "F 主观判级", "老师傅目测判级 {A,B,C}（≠实测）", "老师傅", ["A", "B"]),
    ("rationale", "G 归因", "中文理由：为什么这么调", "老师傅", ["边缘偏低，升边缘温+延时", "沿用基准即可"]),
    ("cause_tag", "G 归因", "成因标签（类目未定→留空）", "老师傅", ["", ""]),
    ("is_ground_truth", "H 复核·一致性", "是否专家真值 TRUE/FALSE（默认 TRUE）", "记录员", ["TRUE", "TRUE"]),
    ("repeat_group_id", "H 复核·一致性", "重复标注组号（隔时重标同一情形用同号）", "记录员", ["", ""]),
    ("measured_quality_grade", "I outcome(出炉后补)", "实测质量判级 {A,B,C}（现在留空）", "后补", ["", ""]),
    ("measured_energy_kwh", "I outcome(出炉后补)", "实测能耗 kWh（现在留空）", "后补", ["", ""]),
]


# 每列的短中文名（说明行与字段字典用；键与 COLUMNS 字段名一一对应）
_ZH: dict[str, str] = {
    "sample_id": "样本号",
    "created_at": "时间",
    "operator_id": "工号",
    "source": "产线/班次",
    "thickness_mm": "厚度mm",
    "glass_type": "品类",
    "quality_mode": "质量模式",
    "x0_95_nm": "X0.95(nm)",
    "iso_t_pct": "IsoT%",
    "ccp_value": "CCP",
    "ccp_is_calibrated": "CCP已标定",
    "stress_image_path": "应力斑图路径",
    "stress_image_sha256": "图像哈希",
    "condition_note": "工况备注",
    "baseline_zone_temps_c": "基准分区温度℃",
    "baseline_zone_roles": "基准分区角色",
    "baseline_temp_upper_c": "基准上炉温℃",
    "baseline_temp_lower_c": "基准下炉温℃",
    "baseline_convection_speed": "基准对流风速",
    "baseline_convection_ratio_upper_lower": "基准对流配比",
    "baseline_oscillation_speed": "基准摆动速度",
    "baseline_oscillation_amplitude": "基准摆动幅度",
    "baseline_heating_duration_s": "基准加热时长s",
    "final_zone_temps_c": "最终分区温度℃",
    "final_zone_roles": "最终分区角色",
    "final_temp_upper_c": "最终上炉温℃",
    "final_temp_lower_c": "最终下炉温℃",
    "final_convection_speed": "最终对流风速",
    "final_convection_ratio_upper_lower": "最终对流配比",
    "final_oscillation_speed": "最终摆动速度",
    "final_oscillation_amplitude": "最终摆动幅度",
    "final_heating_duration_s": "最终加热时长s",
    "expert_quality_grade": "目测判级A/B/C",
    "rationale": "为什么这么调",
    "cause_tag": "成因标签(未定留空)",
    "is_ground_truth": "是否真值",
    "repeat_group_id": "重复组号",
    "measured_quality_grade": "实测判级",
    "measured_energy_kwh": "实测能耗kWh",
}


def _header() -> list[str]:
    """列名表头。"""
    return [c[0] for c in COLUMNS]


def _label_row() -> list[str]:
    """第 2 行中文说明行：★填这里=老师傅列 / 出炉后补 / 勿动。

    sample_id 单元格用“示例-”前缀 → 导入器按既有示例行机制跳过本行，导入零改动。
    """
    cells: list[str] = []
    for name, group, _desc, who, _ex in COLUMNS:
        if name == "sample_id":
            cells.append("示例-说明行(勿删)")
        elif who == "老师傅":
            cells.append(f"★填这里｜{_ZH[name]}")
        elif group.startswith("I"):
            cells.append(f"出炉后补｜{_ZH[name]}")
        else:
            cells.append(f"勿动｜{_ZH[name]}")
    return cells


def _example_rows() -> list[list[str]]:
    """两行格式示例（示意，非真值）。"""
    return [[c[4][i] for c in COLUMNS] for i in range(2)]


def write_csv(path: Path) -> None:
    """写 CSV（utf-8-sig，Excel 打开中文不乱码）：表头 + 中文说明行 + 示例行。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(_header())
        w.writerow(_label_row())
        w.writerows(_example_rows())


def write_xlsx(path: Path) -> bool:
    """写 xlsx（枚举下拉、冻结表头、师傅列高亮、无关列折叠、字段字典页）。openpyxl 不可用则返回 False。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "标注表"
    ws.append(_header())
    ws.append(_label_row())
    for row in _example_rows():
        ws.append(row)
    ws.freeze_panes = "B3"  # 冻结表头两行 + 首列 sample_id

    # 按"谁填"上色：老师傅=黄底加粗、出炉后补=浅蓝、其余=灰底灰字（勿动）
    # 颜色必须写全 8 位 ARGB（FF 不透明）：openpyxl 对 6 位值补 "00" alpha，WPS 会渲染成透明
    fill_master = PatternFill("solid", start_color="FFFFF2CC")
    fill_later = PatternFill("solid", start_color="FFDDEBF7")
    fill_other = PatternFill("solid", start_color="FFE7E6E6")
    for idx, (name, group, _desc, who, _ex) in enumerate(COLUMNS, start=1):
        if who == "老师傅":
            fill, font = fill_master, Font(bold=True)
        elif group.startswith("I"):
            fill, font = fill_later, Font(bold=False)
        else:
            fill, font = fill_other, Font(color="FF808080")
        for r in (1, 2):
            cell = ws.cell(row=r, column=idx)
            cell.fill = fill
            cell.font = font
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = 22 if name in ("final_zone_temps_c", "rationale") else 14

    # 示例行（3–4）灰斜体，一眼区分于真数据
    for r in (3, 4):
        for idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=idx).font = Font(italic=True, color="FF808080")

    # 老师傅不碰的组默认折叠（记录员点列上方"+"展开）：A 余项 / C 输入 / D 基准 / H 复核
    header = _header()

    def _span(first: str, last: str) -> tuple[str, str]:
        """组的起止列字母（按 COLUMNS 顺序定位，列顺序变了自动跟随）。"""
        a = ws.cell(row=1, column=header.index(first) + 1).column_letter
        b = ws.cell(row=1, column=header.index(last) + 1).column_letter
        return a, b

    for first, last in (
        ("created_at", "source"),
        ("x0_95_nm", "condition_note"),
        ("baseline_zone_temps_c", "baseline_heating_duration_s"),
        ("is_ground_truth", "repeat_group_id"),
    ):
        a, b = _span(first, last)
        ws.column_dimensions.group(a, b, hidden=True)
    # OOXML 规范：列带 outlineLevel 时 sheet 级必须声明 outlineLevelCol，缺了 Excel/WPS
    # 按"文件损坏"拒载（2026-07-03 现场踩坑）。openpyxl 从零建簿不自动算；保存器只认
    # column_dimensions.max_outline（并会用它覆盖 sheet_format 上的手动值），故设在这里
    ws.column_dimensions.max_outline = 1

    # 枚举列加下拉校验（数据从第 3 行起；第 2 行是说明行）
    for name, allowed in ENUMS.items():
        col = header.index(name) + 1
        letter = ws.cell(row=1, column=col).column_letter
        dv = DataValidation(type="list", formula1='"' + ",".join(allowed) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}3:{letter}1000")

    # 字段字典页
    ws2 = wb.create_sheet("字段字典")
    ws2.append(["字段", "中文名", "组", "含义/单位/允许值", "谁填"])
    for name, group, desc, who, _ex in COLUMNS:
        ws2.append([name, _ZH[name], group, desc, who])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def main() -> int:
    """生成 CSV 与（可用时）xlsx 到 data/annotation/。"""
    csv_path = OUT_DIR / "标注表模板.csv"
    write_csv(csv_path)
    print(f"已写 {csv_path}（{len(COLUMNS)} 列）")
    xlsx_path = OUT_DIR / "标注表模板.xlsx"
    if write_xlsx(xlsx_path):
        print(f"已写 {xlsx_path}（含下拉/字典页）")
    else:
        print("openpyxl 不可用，跳过 xlsx（仅 CSV）")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except Exception:
        pass
    sys.exit(main())
